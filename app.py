# -*- coding: utf-8 -*-
"""
ThaiSat Freq Checker
โหมด 1: 9.3 (API/A)  — ขอบเขตจาก portal CSV + ความถี่จาก ific.mdb + เทียบฐานไทย
โหมด 2: CR/C (9.7 / 9.12A / 9.21 / 9.14) — tr_aff_ntw (ชื่อข่ายไทยจาก BR) + provn/grp (ย่านความถี่)
ทุกโหมดกรอง wic_no = เลข IFIC ของไฟล์ (อ่านอัตโนมัติ)
ผลลัพธ์มีชีต "สำหรับบันทึก" จัดรูปแบบตามตารางบันทึกเสนอผู้บริหาร (1 แถว/ข่ายต่างชาติ,
ชื่อข่ายไทยเรียงลงบรรทัดในเซลล์เดียว พร้อมก๊อปลง Word)
"""
import streamlit as st
import subprocess, csv, io, tempfile, os
import pandas as pd
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="ThaiSat Freq Checker", layout="wide")

# ---------------- secrets (safe) ----------------
def get_secret(key, default=None):
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default

# ---------------- password gate (robust) ----------------
def check_password():
    pw_secret = get_secret("password")
    if not pw_secret:
        st.warning('ยังไม่ได้ตั้งรหัสผ่านใน Secrets — เพิ่มบรรทัด  password = "รหัสของคุณ"  '
                   'ในหน้า Settings → Secrets เพื่อเปิดด่านรหัสผ่าน (ตอนนี้เข้าใช้ได้โดยไม่ต้องใส่รหัส)')
        return True
    if st.session_state.get("auth_ok"):
        return True
    pw = st.text_input("🔒 รหัสผ่าน", type="password")
    if pw:
        if pw == pw_secret:
            st.session_state["auth_ok"] = True
            return True
        st.error("รหัสผ่านไม่ถูกต้อง")
    return False

# ---------------- mdb helpers ----------------
def mdb_table(mdb_path, table):
    out = subprocess.run(["mdb-export", mdb_path, table],
                         capture_output=True, text=True).stdout
    return list(csv.DictReader(io.StringIO(out)))

def detect_ific(notice_rows):
    vals = [n.get("wic_no", "") for n in notice_rows if n.get("wic_no")]
    return Counter(vals).most_common(1)[0][0] if vals else None

# ---------------- portal CSV ----------------
def read_portal_csv(uploaded):
    raw = uploaded.getvalue().decode("utf-8-sig", errors="replace").splitlines()
    lines = [l for l in raw if not l.lower().startswith("sep=")]
    delim = ";" if (lines and lines[0].count(";") >= lines[0].count(",")) else ","
    return list(csv.DictReader(lines, delimiter=delim))

def col(row, *names):
    for n in names:
        if n in row:
            return row[n]
    return ""

def portal_section_str(r):
    """API/A/14341 หรือ API/A/14116 MOD-1 จากคอลัมน์ portal"""
    s = f"{col(r,'Special section')}/{col(r,'Special section number')}"
    if col(r, "Type of revision") == "M":
        s += f" MOD-{col(r,'Revision number')}"
    return s

# ---------------- formatting ----------------
def fmt_lon(x):
    x = (str(x) if x is not None else "").strip()
    if not x:
        return "NGSO"
    try:
        v = round(float(x), 1)
    except ValueError:
        return x
    return f"{int(v)} E" if v == int(v) else f"{v} E"

def thai_label(sat_name, long_nom):
    """ชื่อข่ายไทยสำหรับบันทึก: GSO -> 'ชื่อ (xx E)', NGSO -> 'ชื่อ (NGSO)'"""
    lon = (long_nom or "").strip()
    return f"{sat_name} ({fmt_lon(lon)})" if lon else f"{sat_name} (NGSO)"

def ssn_map(pub_rows, prefix):
    """ntc_id -> 'CR/C/6245 MOD-2' (เลือกเฉพาะ special section ตาม prefix)"""
    m = {}
    for r in pub_rows:
        if r.get("ssn_ref", "").startswith(prefix):
            s = f"{prefix}/{r['ssn_no']}"
            if r.get("ssn_rev") == "M":
                s += f" MOD-{r.get('ssn_rev_no','')}"
            m[r["ntc_id"]] = s
    return m

# ---------------- Thai base ----------------
def normalise_thai(df):
    df.columns = [c.strip().lower() for c in df.columns]
    need = {"network_name", "freq_start_mhz", "freq_end_mhz"}
    if not need.issubset(df.columns):
        raise ValueError("ฐานไทยต้องมีคอลัมน์: network_name, freq_start_mhz, freq_end_mhz")
    df = df.dropna(subset=["freq_start_mhz", "freq_end_mhz"]).copy()
    df["freq_start_mhz"] = pd.to_numeric(df["freq_start_mhz"], errors="coerce")
    df["freq_end_mhz"] = pd.to_numeric(df["freq_end_mhz"], errors="coerce")
    return df.dropna(subset=["freq_start_mhz", "freq_end_mhz"])

def load_thai(upload, sheet_url):
    if upload is not None:
        df = (pd.read_csv(upload) if upload.name.lower().endswith(".csv")
              else pd.read_excel(upload))
        return normalise_thai(df), "ไฟล์ที่อัปโหลด"
    if sheet_url:
        return normalise_thai(pd.read_csv(sheet_url)), "Google Sheet"
    return None, None

# ---------------- Excel writer (ทั่วไป) ----------------
def build_excel(sheets):
    """sheets = list ของ (ชื่อชีต, DataFrame, [ความกว้างคอลัมน์])"""
    F = "Tahoma"
    hf = Font(name=F, bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F3864")
    bd = Border(*[Side(style="thin", color="D0D0D0")] * 4)
    wb = Workbook(); wb.remove(wb.active)
    for title, df, widths in sheets:
        ws = wb.create_sheet(title[:31])
        ws.append(list(df.columns))
        for _, r in df.iterrows():
            ws.append(list(r))
        for c in ws[1]:
            c.font = hf; c.fill = hfill; c.border = bd
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = Font(name=F); c.border = bd
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        if len(df):
            ws.auto_filter.ref = ws.dimensions
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def nl(items):
    return "\n".join(items) if items else "—"

# ================= โหมด 9.3 (API/A) =================
def build_foreign_apia(mdb_path, portal_rows, ific):
    grp = mdb_table(mdb_path, "grp")
    gbn = defaultdict(list)
    for g in grp:
        if g.get("wic_no") == ific:
            gbn[g["ntc_id"]].append(g)
    apia = [r for r in portal_rows if col(r, "Special section").strip() == "API/A"]
    foreign, nodata = [], []
    for r in apia:
        nid = col(r, "Notice ID"); tgt = col(r, "Targeted notice ID")
        mid = next((x for x in [nid, tgt] if x and x in gbn), None)
        sat = col(r, "Satellite name"); adm = col(r, "Notifying administration")
        if not mid:
            nodata.append((sat, adm, nid, portal_section_str(r)))
            continue
        for g in gbn[mid]:
            if g.get("freq_min"):
                foreign.append({
                    "sat": sat, "adm": adm, "notice_id": mid,
                    "section": portal_section_str(r),
                    "orbit": col(r, "Orbital position") or "NGSO",
                    "beam": g.get("beam_name", ""), "emi": g.get("emi_rcp", ""),
                    "f_min": float(g["freq_min"]), "f_max": float(g["freq_max"]),
                })
    return pd.DataFrame(foreign), apia, nodata

def overlap(foreign, thai, min_khz=0.0):
    # min_khz = 0 -> รวม "แตะขอบ" (0 MHz) ด้วย / >0 -> เฉพาะทับจริงที่กว้างพอ
    min_mhz = min_khz / 1000.0
    rows = []
    for f in foreign.itertuples(index=False):
        for t in thai.itertuples(index=False):
            lo = max(f.f_min, t.freq_start_mhz)
            hi = min(f.f_max, t.freq_end_mhz)
            width = hi - lo
            if width >= min_mhz and width >= 0:
                rows.append({
                    "ดาวเทียมต่างชาติ": f.sat, "adm": f.adm,
                    "ทิศทาง": "ขาลง (E)" if f.emi == "E" else "ขาขึ้น (R)",
                    "ขาลงกระทบภาคพื้น": "✔" if f.emi == "E" else "",
                    "ประเภท": "ทับจริง" if width > 0 else "แตะขอบ (0 MHz)",
                    "beam": f.beam,
                    "ตปท_f_min": round(f.f_min, 4), "ตปท_f_max": round(f.f_max, 4),
                    "ข่ายไทย": t.network_name,
                    "overlap_min": round(lo, 4), "overlap_max": round(hi, 4),
                    "overlap_MHz": round(width, 4),
                })
    return pd.DataFrame(rows)

def memo_93(detail, foreign, nodata):
    """ชีต 'สำหรับบันทึก' — 1 แถว/ข่ายต่างชาติ ตามตารางบันทึก ข้อ 9.3"""
    meta = (foreign.groupby("sat")
            .agg(orbit=("orbit", "first"), section=("section", "first"),
                 adm=("adm", "first"),
                 has_dl=("emi", lambda s: "X" if (s == "E").any() else ""))
            .to_dict("index"))
    rows = []
    for sat, g in detail.groupby("ดาวเทียมต่างชาติ"):
        real = sorted(set(g.loc[g["ประเภท"] == "ทับจริง", "ข่ายไทย"]))
        edge = sorted(set(g.loc[g["ประเภท"] == "แตะขอบ (0 MHz)", "ข่ายไทย"]) - set(real))
        m = meta.get(sat, {})
        rows.append({
            "ข่ายงานดาวเทียม": sat,
            "Nominal Longitude": m.get("orbit", "NGSO"),
            "Special section": m.get("section", ""),
            "ประเทศ": m.get("adm", ""),
            "ข่ายไทยที่กระทบ (S)": nl(real),
            "แตะขอบ (0 MHz)": nl(edge),
            "T (มีขาลง)": m.get("has_dl", ""),
            "มาตรา": "9.3",
        })
    for sat, adm, nid, sec in nodata:
        rows.append({"ข่ายงานดาวเทียม": sat, "Nominal Longitude": "NGSO",
                     "Special section": sec, "ประเทศ": adm,
                     "ข่ายไทยที่กระทบ (S)": "— ไม่มี technical data ในไฟล์รอบนี้ —",
                     "แตะขอบ (0 MHz)": "", "T (มีขาลง)": "", "มาตรา": "9.3"})
    return pd.DataFrame(rows)

# ================= โหมด CR/C =================
def crc_prov_norm(p):
    return "9.7" if p.startswith("9.7") else p

def sec_num(sec):
    """'CR/C/6176 MOD-1' -> '6176'"""
    try:
        return sec.split("/")[-1].split()[0]
    except Exception:
        return ""

def build_crc(mdb_path, ific, thai, allowed_secnums=None, allowed_ids=None):
    """allowed_secnums/ids มาจาก portal CSV (ขอบเขตทางการ) — รายการนอกขอบเขตจะถูกแยกออกมา
    ใน excluded (ไม่ทิ้งเงียบ)"""
    tr = mdb_table(mdb_path, "tr_aff_ntw")
    pv = mdb_table(mdb_path, "provn")
    com = mdb_table(mdb_path, "com_el")
    pub = mdb_table(mdb_path, "pub_ssn")
    grp = mdb_table(mdb_path, "grp")
    cinfo = {c["ntc_id"]: c for c in com}
    ssn = ssn_map(pub, "CR/C")
    gmap = {g["grp_id"]: g for g in grp}

    def in_scope(sec, ids):
        if allowed_secnums is None:
            return True
        return (sec_num(sec) in allowed_secnums) or bool(ids & (allowed_ids or set()))

    # ---- Engine B: 9.7 (รวม -CI/-CII) + 9.12A จาก tr_aff_ntw ----
    agg = defaultdict(lambda: {"thai": set(), "ids": set(), "flags": []})
    for r in tr:
        if r.get("adm") == "THA" and r.get("wic_no") == ific:
            prov = crc_prov_norm(r["coord_prov"])
            sec = ssn.get(r["ntc_id"], f"(ntc {r['ntc_id']})")
            a = agg[(prov, sec)]
            a["ids"].add(r["ntc_id"])
            a["thai"].add(thai_label(r["sat_name"], r.get("long_nom", "")))
            a["flags"].append((r["sat_name"], r.get("f_cause", ""), r.get("f_rec", "")))

    def foreign_of(ids):
        for i in sorted(ids):
            c = cinfo.get(i)
            if c and c.get("sat_name"):
                return c["sat_name"], fmt_lon(c.get("long_nom", "")), c.get("adm", "")
        return f"ID {sorted(ids)[0]}", "?", "?"

    mem_b, flags_rows, excluded = {"9.7": [], "9.12A": []}, [], []
    for (prov, sec), a in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        name, lon, adm = foreign_of(a["ids"])
        if not in_scope(sec, a["ids"]):
            excluded.append({"มาตรา": prov, "Special section": sec, "ข่ายงานดาวเทียม": name,
                             "เหตุผล": "ไม่อยู่ในรายการตีพิมพ์ทางการ (portal) ของรอบนี้"})
            continue
        tgt = mem_b.setdefault(prov, [])
        tgt.append({
            "ข่ายงานดาวเทียม": name, "Nominal Longitude": lon, "BR IFIC": ific,
            "Special section": sec, "ประเทศ": adm,
            "ข่ายงานดาวเทียมที่ได้รับผลกระทบ": nl(sorted(a["thai"])),
            "ข้อบังคับวิทยุ": prov,
        })
        for tname, fc, fr in a["flags"]:
            flags_rows.append({"ข่ายต่างชาติ": name, "มาตรา": prov, "Special section": sec,
                               "ข่ายไทย": tname, "f_cause": fc, "f_rec": fr})

    # ---- Engine A: 9.21/* + 9.14 จาก provn + grp (กรอง wic) ----
    bands = defaultdict(lambda: defaultdict(set))  # (prov, sec) -> (fmin,fmax) -> {emi}
    ids_a = defaultdict(set)
    for p in pv:
        if p.get("adm") == "THA" and (p["coord_prov"].startswith("9.21") or p["coord_prov"] == "9.14"):
            g = gmap.get(p["grp_id"])
            if g and g.get("wic_no") == ific and g.get("freq_min"):
                sec = ssn.get(g["ntc_id"], f"(ntc {g['ntc_id']})")
                key = (p["coord_prov"], sec)
                bands[key][(float(g["freq_min"]), float(g["freq_max"]))].add(g.get("emi_rcp", ""))
                ids_a[key].add(g["ntc_id"])

    mem_a = {"9.21": [], "9.14": []}
    det_a = []
    for (prov, sec), bset in sorted(bands.items(), key=lambda x: (x[0][0], x[0][1])):
        name, lon, adm = foreign_of(ids_a[(prov, sec)])
        if not in_scope(sec, ids_a[(prov, sec)]):
            excluded.append({"มาตรา": prov, "Special section": sec, "ข่ายงานดาวเทียม": name,
                             "เหตุผล": "ไม่อยู่ในรายการตีพิมพ์ทางการ (portal) ของรอบนี้"})
            continue
        real, edge = set(), set()
        if thai is not None:
            for (a, b) in bset:
                for t in thai.itertuples(index=False):
                    lo = max(a, t.freq_start_mhz); hi = min(b, t.freq_end_mhz)
                    if hi - lo > 0:
                        real.add(t.network_name)
                    elif hi - lo == 0:
                        edge.add(t.network_name)
            edge -= real
        band_txt = "\n".join(
            f"{a:g} – {b:g} MHz [{'/'.join(sorted(e for e in emis if e))}]"
            for (a, b), emis in sorted(bset.items()))
        row = {
            "ข่ายงานดาวเทียม": name, "Nominal Longitude": lon, "BR IFIC": ific,
            "Special section": sec, "ประเทศ": adm,
            "ย่านความถี่ (MHz)": band_txt,
            "ข่ายไทยที่ทับ (ทับจริง)": nl(sorted(real)) if thai is not None else "(ต้องมีฐานไทย)",
            "แตะขอบ (0 MHz)": nl(sorted(edge)) if thai is not None else "",
            "มาตรา": prov,
        }
        if prov == "9.14":
            row["หมายเหตุ"] = "ตรวจผลคำนวณ pfd ก่อนแจ้งทักท้วง"
            mem_a["9.14"].append(row)
        else:
            mem_a["9.21"].append(row)
        for (a, b), emis in sorted(bset.items()):
            det_a.append({"ข่ายต่างชาติ": name, "มาตรา": prov, "Special section": sec,
                          "f_min": a, "f_max": b,
                          "ทิศทาง": "/".join(sorted(e for e in emis if e))})

    return (pd.DataFrame(mem_b["9.7"]), pd.DataFrame(mem_b["9.12A"]),
            pd.DataFrame(mem_a["9.21"]), pd.DataFrame(mem_a["9.14"]),
            pd.DataFrame(flags_rows), pd.DataFrame(det_a), pd.DataFrame(excluded))

# ================= UI =================
st.title("🛰️ ตรวจสอบข่ายงานดาวเทียม BR IFIC")

if not check_password():
    st.stop()

mode = st.radio("เลือกโหมด",
                ["ทักท้วง 9.3 (API/A)", "ประสานงาน CR/C (9.7 / 9.12A / 9.21 / 9.14)"],
                horizontal=True)

c1, c2 = st.columns(2)
mdb_file = c1.file_uploader("1) ไฟล์ ific.mdb", type=["mdb"])
portal_file = c2.file_uploader("2) portal CSV (รายการตีพิมพ์ทางการ ITU)", type=["csv"])

sheet_url = get_secret("thai_sheet_csv_url")
if sheet_url:
    st.success("ฐานความถี่ไทย: ใช้จาก **Google Sheet** ที่ตั้งไว้ใน Secrets "
               "(อัปโหลดไฟล์ด้านล่างเพื่อใช้แทนเฉพาะครั้งนี้ได้)")
else:
    st.info("ฐานความถี่ไทย: ยังไม่ได้ตั้ง Google Sheet — อัปโหลดไฟล์ด้านล่าง "
            'หรือเพิ่ม  thai_sheet_csv_url = "ลิงก์ CSV"  ใน Secrets')
thai_file = st.file_uploader("3) ฐานความถี่ไทย (Excel/CSV) — ข้ามได้ถ้าใช้ Google Sheet",
                             type=["xlsx", "csv"])
with st.expander("ℹ️ วิธีตั้ง Google Sheet เป็นฐานความถี่ไทย"):
    st.markdown(
        "1. ใน Google ชีต: **ไฟล์ → แชร์ → เผยแพร่ไปยังเว็บ**\n"
        "2. เลือกรูปแบบ **CSV** → เผยแพร่ → คัดลอกลิงก์ (ลงท้าย `output=csv`)\n"
        "3. ใส่ใน **Settings → Secrets**:  `thai_sheet_csv_url = \"ลิงก์ที่คัดลอก\"`\n"
        "4. ชีตต้องมีคอลัมน์: **network_name, freq_start_mhz, freq_end_mhz**"
    )

is_93 = mode.startswith("ทักท้วง")
if is_93:
    min_khz = st.number_input("กรองจุดทับซ้อนที่เล็กกว่า (kHz) — 0 = เอาทุกจุด (รวมแตะขอบ)",
                              min_value=0.0, value=0.0, step=1.0)

if st.button("▶ ประมวลผล", type="primary"):
    if not mdb_file:
        st.error("ต้องอัปโหลด ific.mdb"); st.stop()
    if is_93 and not portal_file:
        st.error("โหมด 9.3 ต้องอัปโหลด portal CSV (ขอบเขตทางการ)"); st.stop()
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".mdb") as tf:
            tf.write(mdb_file.getvalue()); mdb_path = tf.name

        ific = detect_ific(mdb_table(mdb_path, "notice"))
        try:
            thai, thai_src = load_thai(thai_file, sheet_url)
        except Exception as e:
            st.error(f"อ่านฐานความถี่ไทยไม่สำเร็จ: {e}"); st.stop()

        # ---------- โหมด 9.3 ----------
        if is_93:
            if thai is None:
                st.error("โหมด 9.3 ต้องมีฐานความถี่ไทย"); st.stop()
            portal_rows = read_portal_csv(portal_file)
            foreign, apia, nodata = build_foreign_apia(mdb_path, portal_rows, ific)
            os.unlink(mdb_path)

            st.caption(f"ฐานความถี่ไทย: {thai_src} ({len(thai)} ย่าน)")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("IFIC", ific)
            m2.metric("API/A (ขอบเขต)", len(apia))
            m3.metric("มีความถี่", foreign["sat"].nunique() if len(foreign) else 0)
            m4.metric("ไม่มี tech data", len(nodata))
            if nodata:
                st.warning("ไม่มี technical data (publication-only): "
                           + ", ".join(s for s, _, _, _ in nodata))
            if len(foreign) == 0:
                st.info("ไม่พบความถี่สำหรับ API/A รอบนี้"); st.stop()

            detail = overlap(foreign, thai, min_khz)
            if len(detail) == 0:
                st.success("ไม่มีความถี่ใดทับซ้อนกับฐานไทย"); st.stop()

            n_real = int((detail["ประเภท"] == "ทับจริง").sum())
            n_edge = int((detail["ประเภท"] == "แตะขอบ (0 MHz)").sum())
            st.caption(f"จุดทับซ้อนทั้งหมด {len(detail)} — ทับจริง {n_real} | แตะขอบ (0 MHz) {n_edge}")

            memo = memo_93(detail, foreign, nodata)
            st.subheader("📋 ตารางสำหรับบันทึก (1 แถว/ข่าย — ก๊อปลงบันทึกได้)")
            st.dataframe(memo, width="stretch")
            st.subheader("รายละเอียดจุดทับซ้อน")
            st.dataframe(detail, width="stretch")

            xls = build_excel([
                ("สำหรับบันทึก", memo, [20, 12, 18, 8, 32, 26, 9, 8]),
                ("รายละเอียด", detail, [20, 8, 12, 14, 14, 16, 12, 12, 24, 12, 12, 12]),
            ])
            st.download_button("⬇ ดาวน์โหลด Excel", xls,
                               file_name=f"IFIC{ific}_9.3_ทักท้วง.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # ---------- โหมด CR/C ----------
        else:
            if not portal_file:
                st.error("โหมด CR/C ต้องอัปโหลด portal CSV (ขอบเขตทางการ) — "
                         "บนหน้า portal กด Reset all filters แล้ว Save list as CSV จะได้ครบทุก section ในไฟล์เดียว")
                st.stop()
            prtl = read_portal_csv(portal_file)
            crc_portal = [r for r in prtl if col(r, "Special section").startswith("CR")]
            if not crc_portal:
                st.error("portal CSV นี้ไม่มีรายการ CR/C — ตอน export อย่ากรองเฉพาะ API/A "
                         "(กด Reset all filters แล้ว Save list as CSV)")
                st.stop()
            allowed_secnums = {col(r, "Special section number").strip() for r in crc_portal}
            allowed_ids = set()
            for r in crc_portal:
                for x in (col(r, "Notice ID"), col(r, "Targeted notice ID")):
                    if x:
                        allowed_ids.add(x.strip())
            if thai is None:
                st.warning("ยังไม่มีฐานความถี่ไทย — 9.21/9.14 จะไม่มีคอลัมน์ข่ายไทยที่ทับ")

            m97, m912a, m921, m914, flags, det_a, excluded = build_crc(
                mdb_path, ific, thai, allowed_secnums, allowed_ids)
            os.unlink(mdb_path)

            st.caption(f"ขอบเขต: รายการ CR/C ทางการจาก portal {len(crc_portal)} รายการ")
            if len(excluded):
                st.warning("ข่ายที่พบใน MDB แต่**ไม่อยู่ในรายการตีพิมพ์รอบนี้** — ตัดออกจากผล "
                           "(ตรวจสอบได้ในชีต 'นอกขอบเขต'):")
                st.dataframe(excluded, width="stretch")

            # ข่ายใน portal ที่ไม่ระบุ THA (ปกติ — แค่แจ้งให้ครบ)
            used_nums = set()
            for df in (m97, m912a, m921, m914):
                if len(df):
                    used_nums |= {sec_num(s) for s in df["Special section"]}
            no_tha = [f"{r['Special section']}/{r['Special section number']} {r['Satellite name']}"
                      for r in crc_portal
                      if col(r, "Special section number").strip() not in used_nums]
            if no_tha:
                st.caption("ตีพิมพ์รอบนี้แต่ไม่ระบุ THA (ไม่ต้องดำเนินการ): " + ", ".join(no_tha))

            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("IFIC", ific)
            m2.metric("9.7", len(m97)); m3.metric("9.12A", len(m912a))
            m4.metric("9.21/A-C", len(m921)); m5.metric("9.14", len(m914))

            for title, df in [("มาตรา 9.7 — BR ระบุชื่อข่ายไทย", m97),
                              ("มาตรา 9.12A — BR ระบุชื่อข่ายไทย", m912a),
                              ("มาตรา 9.21/A-C — ย่านที่ระบุ THA + เทียบฐานไทย", m921),
                              ("มาตรา 9.14 — ย่านที่ระบุ THA (สถานีภาคพื้น)", m914)]:
                st.subheader(title)
                if len(df):
                    st.dataframe(df, width="stretch")
                else:
                    st.info("ไม่มีในรอบนี้")

            wb = [("9.7", m97, [20, 12, 8, 18, 8, 34, 10]),
                  ("9.12A", m912a, [20, 12, 8, 18, 8, 34, 10]),
                  ("9.21", m921, [20, 12, 8, 18, 8, 26, 30, 24, 10]),
                  ("9.14", m914, [20, 12, 8, 18, 8, 26, 30, 24, 10, 24]),
                  ("นอกขอบเขต", excluded, [10, 18, 22, 44]),
                  ("flags 9.7-9.12A", flags, [20, 10, 18, 22, 10, 10]),
                  ("รายละเอียดย่าน 9.21-9.14", det_a, [20, 10, 18, 12, 12, 10])]
            xls = build_excel([(t, d, w) for t, d, w in wb if len(d)])
            st.download_button("⬇ ดาวน์โหลด Excel", xls,
                               file_name=f"IFIC{ific}_CRC_9.7-9.14.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"เกิดข้อผิดพลาด: {e}")
